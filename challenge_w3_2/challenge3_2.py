#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# load_wordbook 载入 已有 数据表
# Workbook 处理 新 的数据表
# datetime 处理 时间相关 的数据
from openpyxl import load_workbook,Workbook
import datetime 
### datetime
# .strftime('%Y') 获取年份

### openpyxl
# 文件
# Workbook() 创建文件
### 创建文件默认会创建一个 active 表
# load_workbook('filename.xlsx') 载入 已有 EXCELL 文件 

# 表
# wb.create_sheet('sheetname') 创建 表
# wb.remove('sheetname') 删除 表
# wb.save('filename.xlse')  保存表

# 数据
# sheet.append(list)  添加 行 参数为 列表
# sheet.values 读取所有行 返回列表


wb = load_workbook('courses.xlsx')
# 读取 表 students
students_sheet = wb['students']
# 读取 表 time
time_sheet = wb['time']

#处理 原数据文件 
#合并表格并写入 combine 表中
#保留原数据文件
def combine():
    #创建表 并添加表头
    combine_sheet = wb.create_sheet(title='combine')
    combine_sheet.append(['创建时间','课程名称','学习人数','学习时间'])
    
    for stu in students_sheet.values:
        if stu[2] != '学习时间':
            for time in time_sheet.values:
                if stu[1] == time[1]:
                    combine_sheet.append(list(stu) + [time[2]])
    wb.save('courses.xlsx')


#分割文件
# 读取 combine表 数据
# 将数据安时间分割
# 写入不同的数据表中
def split():
    #读取 combine表
    combine_sheet = wb['combine']
    #获取，所有创建年份 放入 列表
    year_split = []
    for data_line in combine_sheet.values:
        if data_line[0] != '创建时间':
            year_split.append(data_line[0].strftime('%Y'))

    #便利 combine表 数据
    #依据 创建年份 创建文件 添加数据
    for year in set(year_split):
        #创建文件
        wb_teamp = Workbook()
        #删除 已有表 active
        wb_teamp.remove(wb_teamp.active)
        #创建 新表 year
        ws = wb_teamp.create_sheet(title=year)

        #写入数据 
        #筛选数据
        for item_by_year in combine_sheet.values:
            if item_by_year[0] != '创建时间':
                if item_by_year[0].strftime('%Y') == year :
                    ws.append(item_by_year)
        wb_teamp.save('{}.xlsx'.format(year))


    
if __name__ == '__main__':
    combine()
    split()
